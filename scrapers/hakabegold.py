# scrapers/hakabegold.py
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Optional, Tuple
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import requests
import pandas as pd
from openpyxl import load_workbook


URL_HAKABEGOLD = "https://www.logammuliahk.com/#work"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/121.0.0.0 Safari/537.36"
)


@dataclass
class HakabeGoldResult:
    df: pd.DataFrame
    asof_label: str
    buyback_per_gram: Optional[int] = None
    download_url: Optional[str] = None
    iframe_url: Optional[str] = None
    final_iframe_url: Optional[str] = None
    method: Optional[str] = None


def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": UA,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }
    )
    return s


def _fetch_text(s: requests.Session, url: str, timeout: int = 30) -> str:
    r = s.get(url, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    return r.text


def _fetch_bytes(s: requests.Session, url: str, timeout: int = 30) -> bytes:
    headers = {
        "User-Agent": UA,
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream,*/*"
        ),
        "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://www.logammuliahk.com/",
    }
    r = s.get(url, headers=headers, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    return r.content


def _is_xlsx(data: bytes) -> bool:
    return len(data) >= 2 and data[:2] == b"PK"


def _extract_hk_iframe_src(main_html: str) -> str:
    srcs = re.findall(r"<iframe[^>]+src=['\"]([^'\"]+)['\"]", main_html, re.I)
    if not srcs:
        raise RuntimeError("Iframe tidak ditemukan di halaman HK (struktur berubah).")

    for u in srcs:
        lu = u.lower()
        if "1drv.ms/x" in lu or "onedrive.live.com" in lu:
            return u.strip()

    return srcs[0].strip()


def _resolve_final_url(s: requests.Session, url: str) -> str:
    r = s.get(url, timeout=30, allow_redirects=True, stream=True)
    r.raise_for_status()
    return r.url


def _clean_1drv_share(url: str) -> str:
    """
    Buang query embed/office (em=2, ActiveCell, dll).
    Intinya: ambil base shortlink 1drv.ms nya saja.
    """
    p = urlparse(url)
    if "1drv.ms" in p.netloc.lower():
        return urlunparse((p.scheme, p.netloc, p.path, "", "", ""))  # drop query+fragment
    return url


def _ensure_download_1(url: str) -> str:
    """
    Pastikan download=1 ada di query.
    """
    p = urlparse(url)
    qs = parse_qs(p.query)
    qs["download"] = ["1"]
    new_q = urlencode(qs, doseq=True)
    return urlunparse((p.scheme, p.netloc, p.path, p.params, new_q, p.fragment))


def _extract_any_download_link(html: str) -> Optional[str]:
    """
    Kalau response masih HTML, kadang link download disisipkan di dalam HTML.
    Kita cari beberapa pola umum.
    """
    patterns = [
        # onedrive live download with authkey
        r"https://onedrive\.live\.com/download\?[^\"'\s>]+",
        # officecdn download.aspx
        r"https://res\.public\.onecdn\.static\.microsoft[^\"'\s>]+/_layouts/15/download\.aspx[^\"'\s>]+",
        r"https://[^\"'\s>]*officeonline[^\"'\s>]+/_layouts/15/download\.aspx[^\"'\s>]+",
        # kadang ada 1drv.ms direct download link (jarang)
        r"https://1drv\.ms/[^\"'\s>]+\bdownload=1[^\"'\s>]*",
    ]
    for pat in patterns:
        m = re.search(pat, html, re.I)
        if m:
            return m.group(0)
    return None


def _try_download_from_1drv(s: requests.Session, iframe_url: str) -> Tuple[Optional[str], Optional[bytes], str]:
    """
    Strategi utama:
    1) clean 1drv share (drop query)
    2) add ?download=1
    3) GET -> kalau XLSX, done
    4) kalau HTML, cari link download di HTML, fetch lagi
    """
    base = _clean_1drv_share(iframe_url)
    dl = _ensure_download_1(base)

    headers = {
        "User-Agent": UA,
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream,*/*"
        ),
        "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
        "Referer": "https://www.logammuliahk.com/",
    }

    r = s.get(dl, headers=headers, timeout=30, allow_redirects=True)
    r.raise_for_status()

    if _is_xlsx(r.content):
        return r.url, r.content, "1drv_clean_download=1"

    # masih HTML → coba ambil link download dari HTML-nya
    html = r.text if isinstance(r.text, str) else r.content.decode("utf-8", errors="ignore")
    link = _extract_any_download_link(html)
    if link:
        b = _fetch_bytes(s, link)
        if _is_xlsx(b):
            return link, b, "1drv_html_extract_download_link"

    return None, None, "1drv_download_failed"


def _extract_officecdn_download(html: str) -> Optional[str]:
    patterns = [
        r"https://res\.public\.onecdn\.static\.microsoft[^\"'\s>]+/_layouts/15/download\.aspx[^\"'\s>]+",
        r"https://[^\"'\s>]*officeonline[^\"'\s>]+/_layouts/15/download\.aspx[^\"'\s>]+",
    ]
    for p in patterns:
        m = re.search(p, html, re.I)
        if m:
            return m.group(0)
    return None


def _best_effort_download_url_from_params(url: str) -> Optional[str]:
    qs = parse_qs(urlparse(url).query)
    cid = (qs.get("cid") or [None])[0]
    resid = (qs.get("resid") or [None])[0]
    authkey = (qs.get("authkey") or [None])[0]
    if cid and resid and authkey:
        return f"https://onedrive.live.com/download?cid={cid}&resid={resid}&authkey={authkey}"
    return None


def _resolve_download_url(s: requests.Session, iframe_url: str) -> Tuple[str, str, str, bytes]:
    """
    PRIORITAS:
    (A) 1drv.ms clean + download=1 (paling stabil)
    (B) resolve iframe -> HTML -> officecdn download.aspx
    (C) fallback: cari edit? -> build download (authkey)
    """
    # (A)
    dl_url, xlsx_bytes, how = _try_download_from_1drv(s, iframe_url)
    if dl_url and xlsx_bytes:
        final_iframe_url = _resolve_final_url(s, iframe_url)
        return dl_url, how, final_iframe_url, xlsx_bytes

    # (B)
    final_iframe_url = _resolve_final_url(s, iframe_url)
    html = _fetch_text(s, final_iframe_url)

    officecdn = _extract_officecdn_download(html)
    if officecdn:
        xlsx_bytes = _fetch_bytes(s, officecdn)
        if _is_xlsx(xlsx_bytes):
            return officecdn, "officecdn_download_aspx", final_iframe_url, xlsx_bytes

    # (C)
    m = re.search(r"https://onedrive\.live\.com/edit\?[^\"'\s]+", html, re.I)
    if m:
        dl = _best_effort_download_url_from_params(m.group(0))
        if dl:
            xlsx_bytes = _fetch_bytes(s, dl)
            if _is_xlsx(xlsx_bytes):
                return dl, "onedrive_download_with_authkey", final_iframe_url, xlsx_bytes

    raise RuntimeError(
        "Gagal auto-discover link download XLSX.\n"
        f"iframe_url: {iframe_url}\n"
        f"final_iframe_url: {final_iframe_url}\n"
        "Catatan: browser bisa karena JS, tapi server requests sering tidak dapat link itu.\n"
        "Solusi: pastikan iframe src dari Blogger adalah 1drv.ms shortlink dan file OneDrive diset sharing public."
    )


def _to_int(x) -> int:
    if x is None:
        return 0
    if isinstance(x, (int, float)):
        return int(x)
    if isinstance(x, str):
        s = re.sub(r"[^\d]", "", x)
        return int(s) if s.isdigit() else 0
    return 0


def _parse_sheet2_table(xlsx_bytes: bytes) -> HakabeGoldResult:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if str(name).strip().lower() == "sheet2":
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # cari header "Berat"
    start_row = None
    start_col = None
    for r in range(1, 61):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "berat":
                start_row, start_col = r, c
                break
        if start_row:
            break
    if not start_row:
        raise RuntimeError("Tidak menemukan header 'Berat' di XLSX (format berubah).")

    # baca 6 kolom: Berat..Stok
    rows = []
    r = start_row + 1
    while r < start_row + 100:
        w = ws.cell(r, start_col).value
        if w is None or str(w).strip() == "":
            break
        vals = [ws.cell(r, start_col + i).value for i in range(6)]
        rows.append(vals)
        r += 1
    if not rows:
        raise RuntimeError("Tabel harga kosong di XLSX.")

    # buyback per gram (scan bawah)
    buyback = None
    for rr in range(r, min(r + 30, 200)):
        row_text = " ".join([str(ws.cell(rr, cc).value or "") for cc in range(1, 14)]).lower()
        if "buyback" in row_text:
            nums = []
            for cc in range(1, 14):
                v = ws.cell(rr, cc).value
                if isinstance(v, (int, float)):
                    nums.append(int(v))
                elif isinstance(v, str):
                    m = re.sub(r"[^\d]", "", v)
                    if m.isdigit():
                        nums.append(int(m))
            if nums:
                buyback = max(nums)
            break

    # as-of label (ambil string tanggal di header atas)
    asof = "HK Logam Mulia"
    for rr in range(1, min(start_row, 25)):
        for cc in range(1, 14):
            v = ws.cell(rr, cc).value
            if isinstance(v, str) and ("202" in v or "feb" in v.lower() or "jan" in v.lower()):
                asof = v.strip()
                break
        if asof != "HK Logam Mulia":
            break

    out = []
    for w, sell, sell_gr, pph, pph_gr, stok in rows:
        try:
            weight_g = float(w)
        except Exception:
            continue

        sell_idr = _to_int(sell)
        buyback_idr = int(buyback * weight_g) if buyback else 0

        out.append(
            {
                "vendor": "HK Logam Mulia",
                "weight_g": weight_g,
                "sell_idr": sell_idr,
                "buyback_idr": buyback_idr,
                "stock": str(stok or "").strip(),
            }
        )

    df = pd.DataFrame(out)
    return HakabeGoldResult(df=df, asof_label=f"HK Logam Mulia — {asof}", buyback_per_gram=buyback)


def fetch_and_parse_hakabegold() -> HakabeGoldResult:
    s = _session()

    main_html = _fetch_text(s, URL_HAKABEGOLD)
    iframe_url = _extract_hk_iframe_src(main_html)

    download_url, method, final_iframe_url, xlsx_bytes = _resolve_download_url(s, iframe_url)

    if not _is_xlsx(xlsx_bytes):
        head = xlsx_bytes[:200].decode("utf-8", errors="ignore")
        raise RuntimeError(
            "Download tidak mengembalikan XLSX (signature 'PK' tidak ada).\n"
            f"method: {method}\n"
            f"download_url: {download_url}\n"
            f"snippet: {head!r}"
        )

    result = _parse_sheet2_table(xlsx_bytes)
    result.download_url = download_url
    result.iframe_url = iframe_url
    result.final_iframe_url = final_iframe_url
    result.method = method
    return result


def parse_hakabegold(_: str = "") -> Tuple[pd.DataFrame, str]:
    res = fetch_and_parse_hakabegold()
    return res.df, res.asof_label
